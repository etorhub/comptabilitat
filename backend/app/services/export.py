"""Exportacio de dades a CSV, Excel i PDF."""

from __future__ import annotations

import csv
import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Transaction

COLUMNS = [
    ("Data", 12),
    ("Data valor", 12),
    ("Llibre", 16),
    ("Compte", 22),
    ("Concepte", 60),
    ("Comerc", 28),
    ("Categoria", 28),
    ("Import", 14),
    ("Moneda", 8),
    ("Estat", 12),
    ("Etiquetes", 20),
    ("Notes", 30),
]

HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _row(transaction: Transaction) -> list:
    return [
        transaction.booking_date,
        transaction.value_date,
        transaction.ledger.name if transaction.ledger else "",
        transaction.account.display_name if transaction.account else "",
        transaction.description,
        transaction.merchant.display_name if transaction.merchant else "",
        transaction.category.full_name if transaction.category else "",
        transaction.amount,
        transaction.currency,
        transaction.status.value,
        ", ".join(transaction.tags or []),
        transaction.notes,
    ]


def transactions_to_csv(transactions: list[Transaction]) -> bytes:
    """CSV amb separador de punt i coma i BOM, que es el que espera l'Excel."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow([name for name, _ in COLUMNS])
    for transaction in transactions:
        row = _row(transaction)
        writer.writerow(
            [
                value.isoformat()
                if isinstance(value, date)
                else f"{value:.2f}".replace(".", ",")
                if isinstance(value, Decimal)
                else value
                for value in row
            ]
        )
    return buffer.getvalue().encode("utf-8-sig")


def transactions_to_xlsx(transactions: list[Transaction], title: str = "Moviments") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]

    sheet.append([name for name, _ in COLUMNS])
    for index, (_, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width

    for transaction in transactions:
        sheet.append(_row(transaction))

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2):
        for cell in row:
            cell.number_format = "DD/MM/YYYY"
    for row in sheet.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = '#,##0.00 "EUR"'

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def summary_to_xlsx(monthly: list[dict], categories: list[dict], title: str = "Informe") -> bytes:
    """Full amb el resum mensual i el repartiment per categoria."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Mes a mes"

    sheet.append(["Periode", "Ingressos", "Despeses", "Resultat"])
    for row in monthly:
        sheet.append([row["period"], row["income"], row["expenses"], row["net"]])

    breakdown = workbook.create_sheet("Categories")
    breakdown.append(["Categoria", "Import", "Pes", "Moviments"])
    for row in categories:
        breakdown.append(
            [row["category_name"], row["amount"], round(row["share"], 4), row["transactions"]]
        )

    for current in (sheet, breakdown):
        for index in range(1, current.max_column + 1):
            cell = current.cell(row=1, column=index)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            current.column_dimensions[get_column_letter(index)].width = 20
        current.freeze_panes = "A2"

    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = '#,##0.00 "EUR"'
    for row in breakdown.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '#,##0.00 "EUR"'
    for row in breakdown.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "0.0%"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def report_to_pdf(
    *,
    title: str,
    subtitle: str,
    summary: dict[str, Decimal],
    monthly: list[dict],
    categories: list[dict],
) -> bytes:
    """Informe imprimible amb el resum del periode."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    heading = ParagraphStyle(
        "Titol", parent=styles["Title"], fontSize=18, alignment=0, spaceAfter=2
    )
    sub = ParagraphStyle("Subtitol", parent=styles["Normal"], textColor=colors.HexColor("#64748b"))
    section = ParagraphStyle(
        "Seccio", parent=styles["Heading2"], fontSize=12, spaceBefore=14, spaceAfter=6
    )

    def money(value: Decimal) -> str:
        return f"{value:,.2f} EUR".replace(",", " ")

    story: list = [Paragraph(title, heading), Paragraph(subtitle, sub), Spacer(1, 8)]

    story.append(Paragraph("Resum del periode", section))
    story.append(
        _table(
            [["Concepte", "Import"]] + [[name, money(value)] for name, value in summary.items()],
            widths=[110 * mm, 50 * mm],
            align_right={1},
        )
    )

    if monthly:
        story.append(Paragraph("Mes a mes", section))
        story.append(
            _table(
                [["Periode", "Ingressos", "Despeses", "Resultat"]]
                + [
                    [
                        row["period"],
                        money(row["income"]),
                        money(row["expenses"]),
                        money(row["net"]),
                    ]
                    for row in monthly
                ],
                widths=[40 * mm, 40 * mm, 40 * mm, 40 * mm],
                align_right={1, 2, 3},
            )
        )

    if categories:
        story.append(Paragraph("Despeses per categoria", section))
        story.append(
            _table(
                [["Categoria", "Import", "Pes", "Moviments"]]
                + [
                    [
                        row["category_name"],
                        money(row["amount"]),
                        f"{row['share'] * 100:.1f} %",
                        str(row["transactions"]),
                    ]
                    for row in categories
                ],
                widths=[70 * mm, 40 * mm, 25 * mm, 25 * mm],
                align_right={1, 2, 3},
            )
        )

    document.build(story)
    return buffer.getvalue()


def _table(rows: list[list[str]], widths: list[float], align_right: set[int]):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    table = Table(rows, colWidths=widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
    ]
    for column in align_right:
        style.append(("ALIGN", (column, 0), (column, -1), "RIGHT"))
    table.setStyle(TableStyle(style))
    return table
