"""Enmascarar el concepte visible d'un moviment."""

from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal

from sqlalchemy import select

from app.models import Account, BankConnection, Merchant, RecurringSeries, Transaction
from app.models.enums import Cadence, ConnectionStatus, LedgerRole, TransactionStatus
from app.services import reports
from app.services.recurring import detect_recurring
from tests.conftest import grant, login, make_user


def _compte(db, ledger) -> Account:
    connection = BankConnection(
        aspsp_name="Santander", aspsp_country="ES", status=ConnectionStatus.ACTIVE
    )
    db.add(connection)
    db.flush()
    account = Account(
        connection_id=connection.id,
        ledger_id=ledger.id,
        eb_account_uid="uid-mask",
        name="Compte corrent",
    )
    db.add(account)
    db.flush()
    return account


def _moviment(
    db,
    compte,
    *,
    nom="JOAN PUIG",
    amount="-200.00",
    dia=None,
    description=None,
    transfer=None,
    dedup=None,
) -> Transaction:
    ledger_id = compte.ledger_id
    merchant = db.scalar(
        select(Merchant).where(Merchant.ledger_id == ledger_id, Merchant.normalized_name == nom)
    )
    if merchant is None:
        merchant = Merchant(ledger_id=ledger_id, normalized_name=nom, display_name=nom.title())
        db.add(merchant)
        db.flush()
    transaction = Transaction(
        account_id=compte.id,
        ledger_id=ledger_id,
        dedup_key=dedup or f"k-{compte.id}-{nom}-{amount}-{dia or date.today()}",
        booking_date=dia or date.today(),
        amount=Decimal(amount),
        description=description or f"TRANSFERENCIA DE {nom} CONCEPTO: LLOGUER",
        normalized_description=nom,
        counterparty=nom,
        merchant_id=merchant.id,
        status=TransactionStatus.BOOKED,
        transfer_group_id=transfer,
    )
    db.add(transaction)
    db.flush()
    return transaction


def _editor(client, db, ledger):
    user = make_user(db, "editor@example.com")
    grant(db, user, ledger)
    login(client, "editor@example.com")
    return user


def test_enmascarar_amaga_el_concepte_i_el_comerc(client, db, ledgers):
    compte = _compte(db, ledgers["personal"])
    transaccio = _moviment(db, compte)
    _editor(client, db, ledgers["personal"])

    resposta = client.patch(
        f"/api/workspaces/personal/transactions/{transaccio.id}",
        json={"display_description": "Traspàs estalvis"},
    )

    assert resposta.status_code == 200
    body = resposta.json()
    assert body["description"] == "Traspàs estalvis"
    assert body["is_masked"] is True
    assert body["merchant_name"] is None
    assert body["normalized_description"] == ""
    assert body["counterparty"] == ""
    assert "JOAN" not in resposta.text
    assert "display_description" not in body

    llista = client.get("/api/workspaces/personal/transactions").json()["items"]
    assert llista[0]["description"] == "Traspàs estalvis"
    assert llista[0]["merchant_name"] is None
    assert "JOAN" not in str(llista)


def test_buidar_lalias_restaura_el_concepte_del_banc(client, db, ledgers):
    compte = _compte(db, ledgers["personal"])
    transaccio = _moviment(db, compte)
    transaccio.display_description = "Traspàs estalvis"
    db.flush()
    _editor(client, db, ledgers["personal"])

    body = client.patch(
        f"/api/workspaces/personal/transactions/{transaccio.id}",
        json={"display_description": "   "},
    ).json()

    assert body["is_masked"] is False
    assert "JOAN PUIG" in body["description"]
    assert body["merchant_name"] == "Joan Puig"


def test_un_lector_no_pot_enmascarar(client, db, ledgers):
    compte = _compte(db, ledgers["personal"])
    transaccio = _moviment(db, compte)
    user = make_user(db, "sogra@example.com")
    grant(db, user, ledgers["personal"], LedgerRole.VIEWER)
    login(client, "sogra@example.com")

    resposta = client.patch(
        f"/api/workspaces/personal/transactions/{transaccio.id}",
        json={"display_description": "secret"},
    )

    assert resposta.status_code == 403
    db.refresh(transaccio)
    assert transaccio.display_description is None


def test_enmascarar_un_traspas_aplica_les_dues_cames(client, db, ledgers):
    compte = _compte(db, ledgers["personal"])
    sortida = _moviment(db, compte, amount="-500.00", transfer="grup-1", dedup="k-out")
    entrada = _moviment(
        db, compte, amount="500.00", transfer="grup-1", dedup="k-in", description="TRASPAS ENTRADA"
    )
    _editor(client, db, ledgers["personal"])

    client.patch(
        f"/api/workspaces/personal/transactions/{sortida.id}",
        json={"display_description": "Entre comptes"},
    )

    db.refresh(sortida)
    db.refresh(entrada)
    assert sortida.display_description == "Entre comptes"
    assert entrada.display_description == "Entre comptes"

    items = client.get(
        "/api/workspaces/personal/transactions", params={"include_transfers": True}
    ).json()["items"]
    assert {item["description"] for item in items} == {"Entre comptes"}


def test_la_cerca_no_troba_el_text_original_enmascarat(client, db, ledgers):
    compte = _compte(db, ledgers["personal"])
    transaccio = _moviment(db, compte)
    transaccio.display_description = "Traspàs estalvis"
    db.flush()
    _editor(client, db, ledgers["personal"])

    per_nom = client.get("/api/workspaces/personal/transactions", params={"search": "Joan"}).json()
    per_alias = client.get(
        "/api/workspaces/personal/transactions", params={"search": "estalvis"}
    ).json()

    assert per_nom["total"] == 0
    assert per_alias["total"] == 1
    assert per_alias["items"][0]["id"] == transaccio.id


def test_export_csv_i_excel_usen_lalias(client, db, ledgers):
    compte = _compte(db, ledgers["personal"])
    transaccio = _moviment(db, compte)
    transaccio.display_description = "Traspàs estalvis"
    db.flush()
    _editor(client, db, ledgers["personal"])

    csv_text = client.get("/api/workspaces/personal/export/transactions.csv").content.decode(
        "utf-8-sig"
    )
    assert "Traspàs estalvis" in csv_text
    assert "JOAN PUIG" not in csv_text
    assert "TRANSFERENCIA" not in csv_text

    from io import BytesIO

    from openpyxl import load_workbook

    xlsx = client.get("/api/workspaces/personal/export/transactions.xlsx")
    sheet = load_workbook(BytesIO(xlsx.content)).active
    assert sheet.cell(row=2, column=5).value == "Traspàs estalvis"
    assert sheet.cell(row=2, column=6).value in (None, "")


def test_els_informes_de_comerc_ignoren_els_enmascarats(db, ledgers):
    compte = _compte(db, ledgers["personal"])
    visible = _moviment(db, compte, nom="MERCADONA", amount="-30.00", dedup="k-merca")
    amagat = _moviment(db, compte, nom="JOAN PUIG", amount="-200.00", dedup="k-joan")
    amagat.display_description = "Privat"
    db.flush()

    rows = reports.merchant_breakdown(db, [ledgers["personal"].id], None, None)

    noms = {row["merchant_name"] for row in rows}
    assert visible.merchant.display_name in noms
    assert "Joan Puig" not in noms


def test_la_serie_recurrent_agafa_lalias_si_lultima_esta_enmascarada(db, ledgers):
    compte = _compte(db, ledgers["personal"])
    avui = date.today()
    darrer = None
    for index in range(6):
        darrer = _moviment(
            db,
            compte,
            amount="-50.00",
            dia=avui - timedelta(days=30 * (5 - index)),
            dedup=f"k-rec-{index}",
        )
    assert darrer is not None
    darrer.display_description = "Lloguer"
    db.flush()

    stats = detect_recurring(db, ledgers["personal"].id)

    assert stats.created == 1
    series = db.scalar(select(RecurringSeries))
    assert series is not None
    assert series.cadence is Cadence.MONTHLY
    assert series.label == "Lloguer"
