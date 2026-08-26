"""Informes, panell i exportacio."""

from __future__ import annotations

import io
from datetime import date, timedelta
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.time import utcnow
from app.models import Account, Balance, BankConnection, Transaction
from app.models.enums import ConnectionStatus, TransactionStatus
from app.services import reports
from tests.conftest import categoria, grant, login, make_user


@pytest.fixture
def espai(ledgers):
    return ledgers["personal"]


@pytest.fixture
def compte(db, espai) -> Account:
    connection = BankConnection(
        aspsp_name="Santander", aspsp_country="ES", status=ConnectionStatus.ACTIVE
    )
    db.add(connection)
    db.flush()
    account = Account(
        connection_id=connection.id,
        ledger_id=espai.id,
        eb_account_uid="uid-1",
        name="Compte corrent",
    )
    db.add(account)
    db.flush()
    return account


def moviment(db, compte, amount, day, category=None, transfer=None, description="COMPRA"):
    transaction = Transaction(
        account_id=compte.id,
        ledger_id=compte.ledger_id,
        dedup_key=f"k-{description}-{day}-{amount}",
        booking_date=day,
        amount=Decimal(str(amount)),
        description=description,
        normalized_description=description,
        status=TransactionStatus.BOOKED,
        category_id=category.id if category else None,
        transfer_group_id=transfer,
    )
    db.add(transaction)
    db.flush()
    return transaction


def test_els_traspassos_no_compten_com_a_despesa(db, compte, espai):
    avui = date.today()
    moviment(db, compte, "-100.00", avui)
    moviment(db, compte, "-500.00", avui, transfer="grup-1", description="TRASPAS")

    ingressos, despeses = reports.income_and_expenses(db, [espai.id], None, None)

    assert despeses == Decimal("100.00")
    assert ingressos == Decimal("0.00")


def test_els_moviments_exclosos_tampoc_compten(db, compte, espai):
    avui = date.today()
    exclos = moviment(db, compte, "-100.00", avui)
    exclos.is_excluded = True
    db.flush()

    _, despeses = reports.income_and_expenses(db, [espai.id], None, None)

    assert despeses == Decimal("0.00")


def test_el_resum_mensual_separa_ingressos_i_despeses(db, compte, espai):
    avui = date.today().replace(day=15)
    moviment(db, compte, "2000.00", avui, description="NOMINA")
    moviment(db, compte, "-750.00", avui)

    rows = reports.monthly_series(
        db, [espai.id], avui - timedelta(days=40), avui + timedelta(days=1)
    )

    assert len(rows) >= 1
    ultim = rows[-1]
    assert ultim["income"] == Decimal("2000.00")
    assert ultim["expenses"] == Decimal("750.00")
    assert ultim["net"] == Decimal("1250.00")


def test_el_repartiment_agrupa_per_categoria_pare(db, compte, espai, categories):
    avui = date.today()
    supermercat = categoria(db, espai)
    forn = categoria(db, espai, "alimentacio-forn-i-pastisseria")
    moviment(db, compte, "-60.00", avui, category=supermercat)
    moviment(db, compte, "-40.00", avui, category=forn, description="FORN")

    rows = reports.category_breakdown(db, [espai.id], None, None)

    assert len(rows) == 1
    assert rows[0]["category_name"] == "Alimentacio"
    assert rows[0]["amount"] == Decimal("100.00")
    assert rows[0]["share"] == pytest.approx(1.0)
    assert rows[0]["transactions"] == 2


def test_el_panell_suma_els_saldos_dels_llibres_permesos(client, db, compte, espai):
    db.add(
        Balance(
            account_id=compte.id,
            balance_type="CLBD",
            amount=Decimal("1500.00"),
            reference_date=date.today(),
            fetched_at=utcnow(),
        )
    )
    db.flush()
    user = make_user(db, "anna@example.com")
    grant(db, user, espai)
    login(client, "anna@example.com")

    body = client.get("/api/workspaces/personal/analytics/dashboard").json()

    assert Decimal(body["current_balance"]) == Decimal("1500.00")
    assert body["ledger_name"] == "Personal"


def test_exportacio_csv(client, db, compte, espai, categories):
    user = make_user(db, "anna@example.com")
    grant(db, user, espai)
    login(client, "anna@example.com")
    moviment(db, compte, "-42.50", date.today(), description="COMPRA MERCADONA")

    response = client.get("/api/workspaces/personal/export/transactions.csv")

    assert response.status_code == 200
    assert "attachment" in response.headers["content-disposition"]
    text = response.content.decode("utf-8-sig")
    assert "Concepte" in text
    assert "COMPRA MERCADONA" in text
    assert "-42,50" in text, "els decimals han de sortir amb coma per a l'Excel"


def test_exportacio_excel(client, db, compte, espai):
    user = make_user(db, "anna@example.com")
    grant(db, user, espai)
    login(client, "anna@example.com")
    moviment(db, compte, "-42.50", date.today(), description="COMPRA MERCADONA")

    response = client.get("/api/workspaces/personal/export/transactions.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "Data"
    assert sheet.cell(row=2, column=5).value == "COMPRA MERCADONA"
    assert sheet.cell(row=2, column=8).value == -42.5


def test_exportacio_pdf(client, db, compte, espai):
    user = make_user(db, "anna@example.com")
    grant(db, user, espai)
    login(client, "anna@example.com")
    moviment(db, compte, "-42.50", date.today())

    response = client.get("/api/workspaces/personal/export/report.pdf")

    assert response.status_code == 200
    assert response.content.startswith(b"%PDF-")


def test_lexportacio_nomes_porta_dades_de_lespai(client, db, compte, espai, ledgers):
    connection = db.scalar(select(BankConnection))
    altre = Account(
        connection_id=connection.id, ledger_id=ledgers["calella"].id, eb_account_uid="uid-2"
    )
    db.add(altre)
    db.flush()
    moviment(db, compte, "-10.00", date.today(), description="MEU")
    moviment(db, altre, "-99.00", date.today(), description="ALIE")

    user = make_user(db, "anna@example.com")
    grant(db, user, espai)
    login(client, "anna@example.com")

    text = client.get("/api/workspaces/personal/export/transactions.csv").content.decode(
        "utf-8-sig"
    )

    assert "MEU" in text
    assert "ALIE" not in text
